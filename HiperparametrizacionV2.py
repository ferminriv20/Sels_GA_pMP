import openpyxl
import os
import numpy as np
from hyperopt import fmin, tpe, hp, STATUS_OK, Trials
from funtion_pMP import genetic_algorithm_pMP
import pickle

"""
Este script realiza la optimización de los parámetros del algoritmo genético para el problema pMP 
utilizando Hyperopt. Se define un espacio de búsqueda para los hiperparámetros y una función objetivo 
que ejecuta el algoritmo genético con dichos parámetros en varias instancias de prueba, calculando 
una medida de tendencia central del gap obtenido. Los resultados se guardan en un archivo Excel 
y se implementa un mecanismo de checkpoint para guardar el progreso de forma periódica.
"""



TRIALS_FILE = "hyperopt_trials_pmed.pkl"  # archivo donde se guardan los trials
checkpoint = 5  # guardar cada 5 evaluaciones 
counter = 0      # contador de evaluaciones


def save_results(data: dict, name: str):
    '''
    Función para guardar los resultados en un archivo Excel
    '''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    row = 1
    for key, value in data.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value)
        row += 1

    sep = os.sep if os.sep == '/' else '\\'
    directory = f"hyperopt_results{sep}{name}.xlsx"
    workbook.save(directory)
    
def gap(optimo_test, opt):
    return (optimo_test - opt) / opt

def average(*nums: int | float):
    return np.mean(nums)

# Espacio de pruebas {nombre_archivo: (p, optimo)}
test_space = {
    # sencillos
    'pmed39.npy': (10, 9423),
    'pmed9.npy': (40, 2734),
    # medios
    'pmed10.npy': (67, 1255),
    'pmed15.npy': (100, 1729),
    'pmed28.npy': (60, 4498),
    'pmed37.npy': (80, 5057),
    # complicados
    'pmed25.npy': (167, 1828),
    'pmed30.npy': (200, 1989),
    'pmed34.npy': (140, 3013),
    'pmed40.npy': (90, 5128)
}


def objective(params):
    """
    Función objetivo para Hyperopt
    """
    #usar variables globales para checkpoint 
    global counter, trials

    medida_tendencia = params['medida_tendencia']
    iteraciones = int(params['iteraciones'])
    tam = int(params['tam'])
    seleccion = params['seleccion']
    cruzamiento = params['cruzamiento']
    mutacion = params['mutacion']
    prob_mutacion = params['prob_mutacion']
    prob_cruzamiento = params['prob_cruzamiento']
    num_competidores = params['num_competidores']
    sample_frac = params["sample_frac"]
    usar_criterio_parada_cv = params["criterio_parada_cv"]
    frac_mejores = params["frac_mejores"]
    umbral_cv = params["umbral_cv"]
    min_gener = params["min_gener"]
    maximizar = params['maximizar']
    
    # Determinar parámetros para cada operador genético
    if seleccion == 'selecciona_torneo':
        para_seleccion = {'num_competidores': int(num_competidores)}
    elif seleccion == 'ruleta':
        para_seleccion = {}
    else:
        raise ValueError(f"Método de selección {seleccion} no reconocido.")
    
    results = []
    for test in test_space:
        cost_matrix = np.load(f'datasets/{test}')
        facilities = len(cost_matrix)
        p = test_space[test][0]

        if cruzamiento == 'cruzamiento_intercambio':
            para_cruzamiento = {'facilities': facilities}
        else:
            raise ValueError(f"Método de cruzamiento {cruzamiento} no reconocido.")
   
        if mutacion == 'mutacion_simple':
            para_mutacion = {'facilities': facilities}
        else:
            raise ValueError(f"Método de mutación {mutacion} no reconocido.")
    
        # Llamada al algoritmo genético pMP
        optimo = genetic_algorithm_pMP(
            cost_matrix=cost_matrix,
            p=p,
            num_iteraciones=iteraciones,
            pop_size=tam,
            seleccion=seleccion,
            cruzamiento=cruzamiento,
            mutacion=mutacion,
            para_seleccion=para_seleccion,
            para_cruzamiento=para_cruzamiento,
            para_mutacion=para_mutacion,
            prob_cruzamiento=prob_cruzamiento,
            prob_mutacion=prob_mutacion,
            usar_criterio_parada_cv=usar_criterio_parada_cv,
            sample_frac=sample_frac,
            frac_mejores=frac_mejores,
            umbral_cv=umbral_cv,
            min_gener=min_gener,
            maximizar=maximizar
        )[2]
        results.append(gap( test_space[test][1], optimo))
        
    # Medida de tendencia central para normalizar resultados
    best_optimum = medida_tendencia(results)

    #checkpoint periódico
    counter+= 1
    # Guardar estado de trials cada checkpoint evaluaciones
    if (counter % checkpoint == 0) and 'trials' in globals():
        try:
            with open(TRIALS_FILE, "wb") as f:
                pickle.dump(trials, f)
            print(f"[Checkpoint] Guardadas {len(trials.trials)} evaluaciones.")
        except Exception as e:
            print(f"[Checkpoint] Error al guardar trials: {e}")

    return {'loss': best_optimum, 'status': STATUS_OK}


def space(medida_tendencia, iteraciones= 1, tam_max= 200):
    '''
    Define el espacio de búsqueda de hiperparámetros.
    '''
    space = {
        'medida_tendencia': medida_tendencia,
        'iteraciones': iteraciones,
        'tam': hp.quniform('tam', 10, tam_max, 1),
        'seleccion': hp.choice('seleccion', ['selecciona_torneo', 'ruleta']),
        'cruzamiento': hp.choice('cruzamiento', ['cruzamiento_intercambio']),
        'mutacion': hp.choice('mutacion', ['mutacion_simple']),
        'prob_mutacion': hp.uniform('prob_mutacion', 0.1, 0.95),
        'prob_cruzamiento': hp.uniform('prob_cruzamiento', 0.15, 0.975),
        'num_competidores': hp.quniform('num_competidores', 4, 20, 1),
        'sample_frac': 0.37,
        'maximizar': False,
        'criterio_parada_cv': True,
        'frac_mejores': 0.2,
        'umbral_cv': 0.0015,
        'min_gener': 70
    }
    return space



# Número total de evaluaciones
n_evals =  1 #2000
medida = average # Medida de tendencia central a usar

# Cargar trials previos si existen (reanudación)
if os.path.exists(TRIALS_FILE):
    with open(TRIALS_FILE, "rb") as f:
        trials = pickle.load(f)
    print(f"*** Reanudando desde {len(trials.trials)} evaluaciones previas ***")
else:
    trials = Trials()
    print("*** Iniciando nuevo experimento de Hyperopt ***")

# Llamar a fmin usando los trials existentes
try:
    best = fmin(fn=objective, space=space(medida), algo=tpe.suggest, max_evals=n_evals, trials=trials)
    
finally:
    # Guardar siempre los trials al finalizar (sea normal o por error)
    try:
        with open(TRIALS_FILE, "wb") as f:
            pickle.dump(trials, f)
        print(f"[Final] Trials guardados con {len(trials.trials)} evaluaciones.")
    except Exception as e:
        print(f"[Final] Error al guardar trials: {e}")

# Imprimir mejores hiperparámetros y guardarlos en Excel
print("Mejores hiperparámetros:", best)
save_results(best, "hiperparametrizacion_pmed")
